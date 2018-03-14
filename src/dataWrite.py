
import names
import helper
from openpyxl import load_workbook
from copy import copy, deepcopy
from datetime import datetime

import dataRead

dataDir = 'D:\\Chenglin_HR_Work\\01_Payroll\\01_Commission\\2018_Commission\\2018_Commissionstatement\\'

def writeFile(monthIndex, name, department):
    #get month name from index
    month = helper.monthNames[monthIndex-1]
    print("Processing data for " + name+ " for "+month)
    departmentFolder = dataDir + department
    templateFileName = departmentFolder+ '\\Template' + '\\2018commission_'+name+'.xlsx'
    new_file_name = helper.getMonthFolder(departmentFolder,month, monthIndex)+'\\2018commission_'+name+'_'+month+'.xlsx'
    helper.copy_rename(templateFileName, new_file_name)

    wb = load_workbook(filename = new_file_name, data_only=True)

    copyNsData(wb, monthIndex, name)
    fillNsNumOnSummarySheet(wb, monthIndex)

    copyBRData(wb, monthIndex, name, department)
    fillBROnSummarySheet(wb, monthIndex)
    
    wb.save(new_file_name)
    
def copyNsData(wb, monthIndex, name):
    ns_sheet = wb.create_sheet('NS')
    ns_data = dataRead.readNSDataForEmployee(monthIndex, name)
    for row in range(0,len(ns_data)):
        for col in range(0,len(ns_data[row])):
            ns_sheet.cell(row=row+1,column=col+1).value = ns_data[row][col].value
            copyStyle(ns_data[row][col], ns_sheet.cell(row=row+1,column=col+1))

def fillNsNumOnSummarySheet(wb, monthIndex):
    ns_sheet = wb['NS']
    if ns_sheet.max_row == 1:
        return
    
    sum=[0]*monthIndex
    for row in range(2, ns_sheet.max_row+1):
        month = ns_sheet.cell(row=row, column=52).value.month
        sum[month-1] = sum[month-1] + ns_sheet.cell(row=row, column=28).value
    for m in range(1,monthIndex+1):
        fillSummary(wb, sum[m-1], 12, m+1)        

def copyBRData(wb, monthIndex, name, department):
    br_sheet = wb.create_sheet('BR')
    br_data = dataRead.readBRDataForEmployee(monthIndex, name, department)
    #populate data
    for row in range(0,len(br_data)):
        for col in range(0,len(br_data[row])):
            br_sheet.cell(row=row+1,column=col+1).value = br_data[row][col].value
            copyStyle(br_data[row][col], br_sheet.cell(row=row+1,column=col+1))

def fillBROnSummarySheet(wb, monthIndex):
    br_sheet = wb['BR']
    if br_sheet.max_row == 1:
        return

    for m in range(1,monthIndex+1):
        fillSummary(wb, 
            br_sheet.cell(row=br_sheet.max_row,column=10+m).value,17,1+m)

def fillSummary(wb, value, row, col):
    summary = wb['Sales Achievements Details']
    summary.cell(row=row, column=col).value=value

def copyStyle(cell, new_cell):
    if cell.has_style:
        new_cell.font = copy(cell.font)
        new_cell.border = copy(cell.border)
        new_cell.fill = copy(cell.fill)
        new_cell.number_format = copy(cell.number_format)
        new_cell.protection = copy(cell.protection)
        new_cell.alignment = copy(cell.alignment)
