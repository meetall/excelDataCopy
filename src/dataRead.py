from openpyxl import load_workbook
import helper

parentDir = 'D:\\Chenglin_HR_Work\\01_Payroll\\01_Commission\\2018_Commission\\2018_Commissionstatement\\NS&BR_data'
ns_name='NS_data.xlsx'
br_name='BR_data.xlsx'
mnc_br_name='MNC_BRdata.xlsx'

def readNSDataForEmployee(monthIndex, name):
    month = helper.monthNames[monthIndex-1]
    filename = helper.getMonthFolder(parentDir, month, monthIndex)+'\\'+ns_name
    ns = load_workbook(filename = filename, data_only=True)
    orderList = ns['Order_List']

    employeeData=[]
    counter=0
    if orderList == None:
        raise Exception("Could not find the new sales sheet " + filename)
    else:
        if orderList['E1'].value != 'BM':
            raise Exception("Format error, BM is not in the 5th column")
        for row in orderList:
            #copy header first
            if counter == 0:
                employeeData.append(row)
                counter = counter + 1
            if row[4].value !=None and row[4].value.replace(' ','').lower() == name.replace(' ','').lower():
                employeeData.append(row)

    return employeeData

def readBRDataForEmployee(monthIndex, name, department):
    month = helper.monthNames[monthIndex-1]
    if department == 'Carrier' or department == 'Chinese':
        filename = helper.getMonthFolder(parentDir, month, monthIndex)+'\\'+br_name
        return readBRData(monthIndex, name, filename)
    elif department == 'MNC':
        filename = helper.getMonthFolder(parentDir, month, monthIndex)+'\\'+mnc_br_name
        return readBRData(monthIndex, name, filename)
       
def readBRData(monthIndex, name, filename):
    br = load_workbook(filename = filename, data_only=True)
    BR_data = br['BRData']

    employeeData=[]
    counter=0
    endLine = name + "Total"
    if BR_data == None:
        raise Exception("Could not find the BR data sheet " + filename)
    else:
        start = False
        for row in BR_data:    
            #copy header first
            if counter == 0:
                employeeData.append(row)
                counter = counter + 1
            if start == True:
                employeeData.append(row)
                if row[0].value != None:
                    if row[0].value.replace(' ','').lower() == endLine.replace(' ','').lower():
                        break
                    else:
                        raise Exception("Error to copy BR data for "+name+" , could not find end line")
            elif row[0].value != None and row[0].value.replace(' ','').lower() == name.replace(' ','').lower():
                employeeData.append(row)
                start=True

    if start == False:
        print("No data found for " + name)
    return employeeData
                



           